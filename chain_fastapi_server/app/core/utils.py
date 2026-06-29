"""通用工具函数模块。

提供 UUID 生成、文件上传、Excel 解析等功能。
"""

import os
import shutil
import stat
import uuid
from pathlib import Path
from typing import Optional

from fastapi import UploadFile

from app.core.datetime_utils import get_now_str_micro


def get_uuid() -> str:
    """生成 UUID1 并去除连字符，返回 32 位十六进制字符串。"""
    return str(uuid.uuid1()).replace("-", "")


async def upload_xlsx(file: UploadFile, save_path: Optional[str] = None) -> str:
    """上传并保存 xlsx 文件。

    Args:
        file: FastAPI UploadFile 对象。
        save_path: 保存目录，默认为 'uploads/xlsx'。

    Returns:
        保存后的完整文件路径。
    """
    path = save_path or "uploads/xlsx"
    os.makedirs(path, exist_ok=True)

    suffix = Path(file.filename).suffix
    file_full_name = f"{path}/{get_now_str_micro()}_{file.filename}"

    try:
        # 使用临时文件确保写入原子性
        from tempfile import NamedTemporaryFile
        with NamedTemporaryFile(delete=False, suffix=suffix, dir=path) as tmp:
            shutil.copyfileobj(file.file, tmp)
        shutil.move(tmp.name, file_full_name)
        os.chmod(file_full_name, stat.S_IREAD | stat.S_IWRITE | stat.S_IRGRP | stat.S_IROTH)
    finally:
        file.file.close()

    return file_full_name


def parse_xlsx(file_path: str) -> list[dict]:
    """解析 xlsx 文件，返回字典列表。

    第一行作为表头（keys），后续行为数据。

    Args:
        file_path: xlsx 文件路径。

    Returns:
        每行数据的字典列表。

    Raises:
        FileNotFoundError: 文件不存在时抛出。
        ValueError: 文件内容为空时抛出。
    """
    from openpyxl import load_workbook

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        raise ValueError("文件内容为空或仅有表头")

    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        result.append(dict(zip(headers, row)))

    return result
